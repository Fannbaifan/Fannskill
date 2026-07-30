#!/usr/bin/env python3
"""
抖音创作者中心 - 观众画像 + 导出数据抓取 v11
缝合方案：v9的视频列表抓取 + v10的URL直跳 + DOM正则提取
"""
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
import json, time, re, os

USER_DATA_DIR = r"D:\HuaweiMoveData\Users\Fann\AppData\Local\Microsoft\Edge\User Data"
DOWNLOAD_DIR = r"D:\HuaweiMoveData\Users\Fann\Downloads"


def safe_goto(page, url, timeout=60000):
    for attempt in range(3):
        try:
            page.goto(url, timeout=timeout, wait_until="domcontentloaded")
            page.wait_for_timeout(3000)
            return True
        except PlaywrightTimeout:
            time.sleep(2)
    return False


def parse_excel(filepath):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))
        return {
            "file": os.path.basename(filepath),
            "total_play": int(sum(r.get('播放量', 0) or 0 for r in data)),
            "total_like": int(sum(r.get('点赞量', 0) or 0 for r in data)),
            "total_comment": int(sum(r.get('评论量', 0) or 0 for r in data)),
            "total_share": int(sum(r.get('分享量', 0) or 0 for r in data)),
            "total_collect": int(sum(r.get('收藏量', 0) or 0 for r in data)),
            "items_count": len(data)
        }
    except Exception as e:
        return {"file": os.path.basename(filepath), "error": str(e)}


def click_export(page):
    print("  点击导出数据...")
    try:
        with page.expect_download(timeout=30000) as download_info:
            page.click('text=导出数据', timeout=5000)
        download = download_info.value
        filepath = os.path.join(DOWNLOAD_DIR, "douyin_export.xlsx")
        download.save_as(filepath)
        print(f"  ✓ 已保存: {filepath}")
        return parse_excel(filepath)
    except Exception as e:
        print(f"  ⚠ 导出失败: {e}")
        return None


def get_videos_from_page_v9(page):
    """v9 的视频列表抓取方法（从页面文本中找包含 # 的内容）"""
    print("  从页面抓取视频...")
    
    all_texts = page.evaluate('''() => {
        return Array.from(document.querySelectorAll('body *'))
            .map(el => el.textContent.trim())
            .filter(t => t.length > 5 && t.length < 200);
    }''')
    
    videos = []
    seen = set()
    for text in all_texts:
        if '#' in text and text not in seen and len(videos) < 30:
            seen.add(text)
            videos.append({"title": text[:100], "index": len(videos)})
    
    return videos


def get_item_ids_from_page(page):
    """从当前页面提取所有 item_id"""
    print("  提取 item_id...")
    
    # 等待一下确保页面加载完
    page.wait_for_timeout(2000)
    
    items = page.evaluate('''() => {
        const result = [];
        const seen = new Set();
        
        // 方法1: 从 a 标签 href
        document.querySelectorAll('a[href*="item_id="]').forEach(a => {
            const m = a.href.match(/item_id=(\\d+)/);
            if (m && !seen.has(m[1])) {
                seen.add(m[1]);
                result.push({item_id: m[1], title: a.textContent.trim().slice(0,80)});
            }
        });
        
        // 方法2: 从 data-item-id
        if (result.length === 0) {
            document.querySelectorAll('[data-item-id]').forEach(el => {
                const id = el.getAttribute('data-item-id');
                if (id && !seen.has(id)) {
                    seen.add(id);
                    result.push({item_id: id, title: el.textContent.trim().slice(0,80)});
                }
            });
        }
        
        return result;
    }''')
    
    return items


def extract_audience_from_body(page):
    """从页面 body 文本正则提取观众画像"""
    body_text = page.inner_text("body")
    
    gender = {}
    age = {}
    region = {}
    
    for m in re.finditer(r'(男性|女性|男|女)\s*(\d+(?:\.\d+)?)\s*%', body_text):
        key = 'male' if m.group(1) in ('男性', '男') else 'female'
        gender[key] = float(m.group(2))
    
    for m in re.finditer(r'(\d+[-+]\d+|\d+\+)\s*(?:岁)?\s*(\d+(?:\.\d+)?)\s*%', body_text):
        age[m.group(1)] = float(m.group(2))
    
    provinces = ['北京','上海','天津','重庆','河北','山西','辽宁','吉林','黑龙江',
                 '江苏','浙江','安徽','福建','江西','山东','河南','湖北','湖南',
                 '广东','海南','四川','贵州','云南','陕西','甘肃','青海','台湾',
                 '内蒙古','广西','西藏','宁夏','新疆','香港','澳门']
    for prov in provinces:
        for m in re.finditer(rf'{prov}\s*(\d+(?:\.\d+)?)\s*%', body_text):
            region[prov] = float(m.group(1))
    
    return {"gender": gender, "age": age, "region": dict(sorted(region.items(), key=lambda x: -x[1])[:5])}


def fetch_audience_for_video(page, item_id, title):
    """抓取单条视频的观众画像"""
    try:
        print(f"    访问观众分析页面...")
        
        # 直接构造 URL 访问
        audience_url = f"https://creator.douyin.com/creator-micro/data-center/content/audience?item_id={item_id}"
        if not safe_goto(page, audience_url, timeout=45000):
            print(f"    ❌ 无法打开页面")
            return None
        
        # 强制等待 4 秒
        print(f"    等待渲染 (4秒)...")
        page.wait_for_timeout(4000)
        
        # 尝试点击「观众分析」标签
        print(f"    点击观众分析标签...")
        clicked = False
        
        try:
            page.locator('div[role="tab"]').get_by_text('观众分析', exact=True).first.click(timeout=5000)
            clicked = True
            print(f"    ✓ 方法A")
        except Exception as e:
            print(f"    方法A: {e}")
        
        if not clicked:
            try:
                page.locator("//div[contains(text(), '观众分析')]").first.click(timeout=5000)
                clicked = True
                print(f"    ✓ 方法B")
            except Exception as e:
                print(f"    方法B: {e}")
        
        if not clicked:
            try:
                page.evaluate('''() => {
                    const els = Array.from(document.querySelectorAll('*'));
                    const target = els.find(el => el.textContent.trim() === '观众分析');
                    if (target) target.click();
                }''')
                clicked = True
                print(f"    ✓ 方法C")
            except Exception as e:
                print(f"    方法C: {e}")
        
        if not clicked:
            print(f"    ❌ 所有方法失败。URL: {page.url}")
            return None
        
        # 等待 + 滚动
        page.wait_for_timeout(4000)
        page.evaluate("window.scrollBy(0, 800)")
        page.wait_for_timeout(2000)
        page.evaluate("window.scrollBy(0, 800)")
        page.wait_for_timeout(2000)
        
        # 提取数据
        profile = extract_audience_from_body(page)
        
        if not profile['gender'] and not profile['age']:
            print(f"    该视频暂无观众数据")
            return None
        
        return profile
    
    except Exception as e:
        print(f"    ❌ 异常: {e}")
        try:
            print(f"    URL: {page.url}")
        except:
            pass
        return None


def main():
    print("=" * 50)
    print("抖音数据抓取 v11")
    print("=" * 50)
    
    with sync_playwright() as p:
        print("\n启动 Edge...")
        browser = p.chromium.launch_persistent_context(
            USER_DATA_DIR,
            channel="msedge",
            headless=False,
            args=["--no-sandbox", "--disable-setuid-sandbox"],
            viewport={"width": 1440, "height": 900}
        )
        page = browser.new_page()
        
        # 登录检查
        print("\n检查登录...")
        if not safe_goto(page, "https://creator.douyin.com/"):
            print("❌ 无法打开")
            browser.close()
            return
        
        if "login" in page.url.lower():
            print("⚠️ 请扫码登录，按回车继续")
            input()
            page.reload()
            page.wait_for_timeout(3000)
        
        print("✅ 已登录")
        
        # 进入作品分析
        print("\n进入作品分析...")
        safe_goto(page, "https://creator.douyin.com/creator-micro/data-center/content")
        
        # 点击投稿列表
        print("\n点击投稿列表...")
        for _ in range(10):
            try:
                page.click('text=投稿列表', timeout=2000)
                print("  ✓ 已点击")
                break
            except:
                time.sleep(1)
        
        page.wait_for_timeout(4000)
        
        # 导出数据
        print("\n导出数据...")
        export_data = click_export(page)
        if export_data:
            print(f"  总播放: {export_data.get('total_play', 0):,}")
        
        # ========== 获取视频列表（v9方式）==========
        print("\n获取视频列表...")
        videos_text = get_videos_from_page_v9(page)
        print(f"  文本抓取: {len(videos_text)} 条")
        
        # 同时尝试提取 item_id
        items_with_id = get_item_ids_from_page(page)
        print(f"  item_id提取: {len(items_with_id)} 条")
        
        # 合并：优先用带 item_id 的，如果不够再用文本抓取的
        videos = []
        if items_with_id:
            videos = items_with_id
        else:
            # 如果没有 item_id，用文本列表（但无法抓画像）
            videos = [{"item_id": "", "title": v["title"]} for v in videos_text]
        
        print(f"  最终: {len(videos)} 条")
        
        # 抓取观众画像
        print("\n抓取观众画像...")
        results = []
        for i, v in enumerate(videos[:15]):
            print(f"\n[{i+1}/{min(len(videos),15)}] {v['title'][:50]}")
            
            if not v.get('item_id'):
                print("   无 item_id，跳过")
                continue
            
            profile = fetch_audience_for_video(page, v['item_id'], v['title'])
            if profile:
                results.append({
                    "item_id": v['item_id'],
                    "title": v['title'],
                    "audience": profile
                })
                print(f"   性别: {profile['gender']}")
                print(f"   年龄: {profile['age']}")
                print(f"   地域: {profile['region']}")
            else:
                print("   跳过")
        
        browser.close()
    
    # 保存
    output = {
        "fetch_time": time.strftime("%Y-%m-%d %H:%M:%S"),
        "export_data": export_data,
        "total_videos": len(videos),
        "audience_count": len(results),
        "videos": results
    }
    
    with open("douyin_audience_data.json", "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f"\n{'='*50}")
    print(f"✅ 完成！")
    print(f"   导出: {'✓' if export_data else '✗'}")
    print(f"   画像: {len(results)} 条")
    print(f"   文件: douyin_audience_data.json")


if __name__ == '__main__':
    main()
