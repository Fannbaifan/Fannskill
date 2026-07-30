#!/usr/bin/env python3
"""
抖音创作者中心 - 观众画像 + 导出数据抓取 v9
修复：去掉pandas、精准点击观众分析、全文正则匹配、try...except防崩溃
"""
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
import json, time, re, os

USER_DATA_DIR = r"D:\HuaweiMoveData\Users\Fann\AppData\Local\Microsoft\Edge\User Data"
DOWNLOAD_DIR = r"D:\HuaweiMoveData\Users\Fann\Downloads"


def safe_goto(page, url, timeout=60000):
    for attempt in range(3):
        try:
            page.goto(url, timeout=timeout, wait_until="domcontentloaded")
            page.wait_for_timeout(3000)
            return True
        except PlaywrightTimeout:
            time.sleep(2)
    return False


def parse_excel_without_pandas(filepath):
    """用 openpyxl 读取 Excel，无需 pandas"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(dict(zip(headers, row)))
        
        total_play = sum(r.get('播放量', 0) or 0 for r in data)
        total_like = sum(r.get('点赞量', 0) or 0 for r in data)
        total_comment = sum(r.get('评论量', 0) or 0 for r in data)
        total_share = sum(r.get('分享量', 0) or 0 for r in data)
        total_collect = sum(r.get('收藏量', 0) or 0 for r in data)
        
        return {
            "file": os.path.basename(filepath),
            "total_play": int(total_play),
            "total_like": int(total_like),
            "total_comment": int(total_comment),
            "total_share": int(total_share),
            "total_collect": int(total_collect),
            "items_count": len(data)
        }
    except Exception as e:
        return {"file": os.path.basename(filepath), "error": str(e)}


def click_export_with_download(page):
    """使用 expect_download 点击导出并保存文件"""
    print("  点击导出数据...")
    
    try:
        with page.expect_download(timeout=30000) as download_info:
            page.click('text=导出数据', timeout=5000)
        
        download = download_info.value
        filepath = os.path.join(DOWNLOAD_DIR, "douyin_export.xlsx")
        download.save_as(filepath)
        print(f"  ✓ 已保存: {filepath}")
        
        return parse_excel_without_pandas(filepath)
    
    except Exception as e:
        print(f"  ⚠ 导出失败: {e}")
        return None


def get_videos_from_page(page):
    """从页面文本抓取视频列表"""
    print("  从页面抓取视频...")
    
    all_texts = page.evaluate('''() => {
        return Array.from(document.querySelectorAll('body *'))
            .map(el => el.textContent.trim())
            .filter(t => t.length > 5 && t.length < 200);
    }''')
    
    videos = []
    seen = set()
    for text in all_texts:
        if '#' in text and text not in seen and len(videos) < 25:
            seen.add(text)
            videos.append({"title": text[:100], "index": len(videos)})
    
    return videos


def extract_audience_from_body_text(page):
    """用 inner_text('body') + 正则匹配提取观众画像"""
    # 获取页面全部可见文本
    body_text = page.inner_text("body")
    
    gender = {}
    age = {}
    region = {}
    
    # 性别：匹配 "男性 86%" 或 "男 86%"
    for m in re.finditer(r'(男性|女性|男|女)\s*(\d+(?:\.\d+)?)\s*%', body_text):
        key = 'male' if m.group(1) in ('男性', '男') else 'female'
        gender[key] = float(m.group(2))
    
    # 年龄：匹配 "18-23 45%" 或 "24-30岁 30%"
    for m in re.finditer(r'(\d+[-+]\d+|\d+\+)\s*(?:岁)?\s*(\d+(?:\.\d+)?)\s*%', body_text):
        age[m.group(1)] = float(m.group(2))
    
    # 地域：匹配 "广东 11.90%"
    provinces = ['北京','上海','天津','重庆','河北','山西','辽宁','吉林','黑龙江',
                 '江苏','浙江','安徽','福建','江西','山东','河南','湖北','湖南',
                 '广东','海南','四川','贵州','云南','陕西','甘肃','青海','台湾',
                 '内蒙古','广西','西藏','宁夏','新疆','香港','澳门']
    for prov in provinces:
        for m in re.finditer(rf'{prov}\s*(\d+(?:\.\d+)?)\s*%', body_text):
            region[prov] = float(m.group(1))
    
    return {
        "gender": gender,
        "age": age,
        "region": dict(sorted(region.items(), key=lambda x: -x[1])[:5])
    }


def fetch_single_audience(page, video_index):
    """抓取单条视频的观众画像，全程 try...except"""
    try:
        # 回到投稿列表
        if not safe_goto(page, "https://creator.douyin.com/creator-micro/data-center/content"):
            return None
        
        # 点击投稿列表标签
        for _ in range(5):
            try:
                page.click('text=投稿列表', timeout=2000)
                break
            except:
                time.sleep(1)
        
        page.wait_for_timeout(3000)
        
        # 点击第 video_index 条视频的「分析详情」
        detail_links = page.query_selector_all('text=分析详情')
        if video_index >= len(detail_links):
            print("    未找到分析详情按钮")
            return None
        
        detail_links[video_index].click()
        page.wait_for_timeout(3000)
        
        # 点击「观众分析」标签
        print("    点击观众分析...")
        clicked = False
        for _ in range(5):
            try:
                page.get_by_text("观众分析", exact=True).click(timeout=3000)
                clicked = True
                print("    ✓ 已点击")
                break
            except:
                try:
                    page.locator('div[role="tab"]:has-text("观众分析")').click(timeout=3000)
                    clicked = True
                    print("    ✓ 已点击(tab)")
                    break
                except:
                    time.sleep(1)
        
        if not clicked:
            print("    ⚠ 未找到观众分析标签")
            return None
        
        # 强制等待 4 秒
        print("    等待数据加载 (4秒)...")
        page.wait_for_timeout(4000)
        
        # 向下滚动触发图表渲染
        print("    滚动触发图表...")
        page.evaluate("window.scrollBy(0, 800)")
        page.wait_for_timeout(2000)
        page.evaluate("window.scrollBy(0, 800)")
        page.wait_for_timeout(2000)
        
        # 用全文正则匹配提取数据
        profile = extract_audience_from_body_text(page)
        
        # 检查是否真的有数据
        if not profile['gender'] and not profile['age']:
            print("    该视频暂无观众数据")
            return None
        
        return profile
    
    except Exception as e:
        print(f"    错误: {e}")
        return None


def main():
    print("=" * 50)
    print("抖音数据抓取 v9")
    print("=" * 50)
    
    with sync_playwright() as p:
        print("\n启动 Edge...")
        browser = p.chromium.launch_persistent_context(
            USER_DATA_DIR,
            channel="msedge",
            headless=False,
            args=["--no-sandbox", "--disable-setuid-sandbox"],
            viewport={"width": 1440, "height": 900}
        )
        page = browser.new_page()
        
        # 登录检查
        print("\n检查登录...")
        if not safe_goto(page, "https://creator.douyin.com/"):
            print("❌ 无法打开")
            browser.close()
            return
        
        if "login" in page.url.lower():
            print("⚠️ 请扫码登录，按回车继续")
            input()
            page.reload()
            page.wait_for_timeout(3000)
        
        print("✅ 已登录")
        
        # 进入作品分析
        print("\n进入作品分析...")
        safe_goto(page, "https://creator.douyin.com/creator-micro/data-center/content")
        
        # 点击投稿列表
        print("\n点击投稿列表...")
        for _ in range(10):
            try:
                page.click('text=投稿列表', timeout=2000)
                print("  ✓ 已点击")
                break
            except:
                time.sleep(1)
        
        page.wait_for_timeout(4000)
        
        # 导出数据
        print("\n导出数据...")
        export_data = click_export_with_download(page)
        if export_data:
            print(f"  总播放: {export_data.get('total_play', 0):,}")
            print(f"  总点赞: {export_data.get('total_like', 0):,}")
        
        # 获取视频列表
        print("\n获取视频列表...")
        videos = get_videos_from_page(page)
        print(f"  找到 {len(videos)} 条")
        
        # 抓取观众画像
        print("\n抓取观众画像...")
        results = []
        for i, v in enumerate(videos[:15]):
            print(f"\n[{i+1}/{min(len(videos),15)}] {v['title'][:50]}")
            
            profile = fetch_single_audience(page, i)
            if profile:
                results.append({"title": v['title'], "audience": profile})
                print(f"   性别: {profile['gender']}")
                print(f"   年龄: {profile['age']}")
                print(f"   地域: {profile['region']}")
            else:
                print("   跳过（无数据或失败）")
        
        browser.close()
    
    # 保存
    output = {
        "fetch_time": time.strftime("%Y-%m-%d %H:%M:%S"),
        "export_data": export_data,
        "total_videos": len(videos),
        "audience_count": len(results),
        "videos": results
    }
    
    with open("douyin_audience_data.json", "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f"\n{'='*50}")
    print(f"✅ 完成！")
    print(f"   导出: {'✓' if export_data else '✗'}")
    print(f"   画像: {len(results)} 条")
    print(f"   文件: douyin_audience_data.json")


if __name__ == '__main__':
    main()
