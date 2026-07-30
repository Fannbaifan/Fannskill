#!/usr/bin/env python3
"""
抖音创作者中心 - 观众画像 + 导出数据抓取 v12
从已下载的 Excel 中提取 item_id，不再去网页 DOM 里找
"""
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
import json, time, re, os

USER_DATA_DIR = r"D:\HuaweiMoveData\Users\Fann\AppData\Local\Microsoft\Edge\User Data"
DOWNLOAD_DIR = r"D:\HuaweiMoveData\Users\Fann\Downloads"
EXPORT_FILE = os.path.join(DOWNLOAD_DIR, "douyin_export.xlsx")


def safe_goto(page, url, timeout=60000):
    for attempt in range(3):
        try:
            page.goto(url, timeout=timeout, wait_until="domcontentloaded")
            page.wait_for_timeout(3000)
            return True
        except PlaywrightTimeout:
            time.sleep(2)
    return False


def click_export(page):
    """点击导出并下载"""
    print("  点击导出数据...")
    try:
        with page.expect_download(timeout=30000) as download_info:
            page.click('text=导出数据', timeout=5000)
        download = download_info.value
        download.save_as(EXPORT_FILE)
        print(f"  ✓ 已保存: {EXPORT_FILE}")
        return True
    except Exception as e:
        print(f"  ⚠ 导出失败: {e}")
        return False


def parse_export_excel():
    """从已下载的 Excel 中提取 item_id 和标题"""
    print("  解析 Excel...")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXPORT_FILE, data_only=True)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        print(f"    列名: {headers}")
        
        # 找到关键列的索引
        title_col = None
        link_col = None
        id_col = None
        play_col = None
        like_col = None
        
        for i, h in enumerate(headers):
            if h in ('作品名称', '标题', 'description'):
                title_col = i
            elif h in ('作品链接', '链接', 'link'):
                link_col = i
            elif h in ('作品ID', 'item_id', '视频ID'):
                id_col = i
            elif h == '播放量':
                play_col = i
            elif h == '点赞量':
                like_col = i
        
        videos = []
        total_play = 0
        total_like = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            
            title = row[title_col] if title_col is not None and title_col < len(row) else ''
            link = row[link_col] if link_col is not None and link_col < len(row) else ''
            item_id = row[id_col] if id_col is not None and id_col < len(row) else ''
            
            # 如果 id_col 没找到，从链接中提取
            if not item_id and link:
                m = re.search(r'item_id=(\d+)', str(link))
                if m:
                    item_id = m.group(1)
            
            # 如果还是没找到，尝试从其他列找数字ID
            if not item_id:
                for val in row:
                    if isinstance(val, (int, float)) and val > 1000000000000:
                        item_id = str(int(val))
                        break
            
            if item_id and title:
                videos.append({
                    "item_id": str(item_id),
                    "title": str(title)[:100]
                })
            
            if play_col is not None and play_col < len(row):
                total_play += row[play_col] or 0
            if like_col is not None and like_col < len(row):
                total_like += row[like_col] or 0
        
        print(f"    提取到 {len(videos)} 条视频")
        return {
            "videos": videos,
            "total_play": int(total_play),
            "total_like": int(total_like),
            "items_count": len(videos)
        }
    
    except Exception as e:
        print(f"    ❌ 解析失败: {e}")
        return {"videos": [], "total_play": 0, "total_like": 0, "items_count": 0, "error": str(e)}


def extract_audience_from_body(page):
    """从页面 body 文本正则提取观众画像"""
    body_text = page.inner_text("body")
    
    gender = {}
    age = {}
    region = {}
    
    for m in re.finditer(r'(男性|女性|男|女)\s*(\d+(?:\.\d+)?)\s*%', body_text):
        key = 'male' if m.group(1) in ('男性', '男') else 'female'
        gender[key] = float(m.group(2))
    
    for m in re.finditer(r'(\d+[-+]\d+|\d+\+)\s*(?:岁)?\s*(\d+(?:\.\d+)?)\s*%', body_text):
        age[m.group(1)] = float(m.group(2))
    
    provinces = ['北京','上海','天津','重庆','河北','山西','辽宁','吉林','黑龙江',
                 '江苏','浙江','安徽','福建','江西','山东','河南','湖北','湖南',
                 '广东','海南','四川','贵州','云南','陕西','甘肃','青海','台湾',
                 '内蒙古','广西','西藏','宁夏','新疆','香港','澳门']
    for prov in provinces:
        for m in re.finditer(rf'{prov}\s*(\d+(?:\.\d+)?)\s*%', body_text):
            region[prov] = float(m.group(1))
    
    return {"gender": gender, "age": age, "region": dict(sorted(region.items(), key=lambda x: -x[1])[:5])}


def fetch_audience_for_video(page, item_id, title):
    """抓取单条视频的观众画像"""
    try:
        print(f"    访问观众分析...")
        
        # 直接构造 URL
        audience_url = f"https://creator.douyin.com/creator-micro/data-center/content/audience?item_id={item_id}"
        if not safe_goto(page, audience_url, timeout=45000):
            print(f"    ❌ 无法打开")
            return None
        
        # 强制等待
        page.wait_for_timeout(4000)
        
        # 尝试点击「观众分析」标签
        print(f"    点击标签...")
        clicked = False
        
        for method_name, click_fn in [
            ("A", lambda: page.locator('div[role="tab"]').get_by_text('观众分析', exact=True).first.click(timeout=5000)),
            ("B", lambda: page.locator("//div[contains(text(), '观众分析')]").first.click(timeout=5000)),
            ("C", lambda: page.evaluate('''() => { const els = Array.from(document.querySelectorAll('*')); const t = els.find(el => el.textContent.trim() === '观众分析'); if (t) t.click(); }''')),
        ]:
            try:
                click_fn()
                clicked = True
                print(f"    ✓ 方法{method_name}")
                break
            except Exception as e:
                print(f"    方法{method_name}: {str(e)[:50]}")
        
        if not clicked:
            print(f"    ❌ 所有方法失败。URL: {page.url}")
            return None
        
        # 等待 + 滚动
        page.wait_for_timeout(4000)
        page.evaluate("window.scrollBy(0, 800)")
        page.wait_for_timeout(2000)
        page.evaluate("window.scrollBy(0, 800)")
        page.wait_for_timeout(2000)
        
        # 提取数据
        profile = extract_audience_from_body(page)
        
        if not profile['gender'] and not profile['age']:
            print(f"    暂无观众数据")
            return None
        
        return profile
    
    except Exception as e:
        print(f"    ❌ 异常: {e}")
        try:
            print(f"    URL: {page.url}")
        except:
            pass
        return None


def main():
    print("=" * 50)
    print("抖音数据抓取 v12")
    print("=" * 50)
    
    with sync_playwright() as p:
        print("\n启动 Edge...")
        browser = p.chromium.launch_persistent_context(
            USER_DATA_DIR,
            channel="msedge",
            headless=False,
            args=["--no-sandbox", "--disable-setuid-sandbox"],
            viewport={"width": 1440, "height": 900}
        )
        page = browser.new_page()
        
        # 登录检查
        print("\n检查登录...")
        if not safe_goto(page, "https://creator.douyin.com/"):
            print("❌ 无法打开")
            browser.close()
            return
        
        if "login" in page.url.lower():
            print("⚠️ 请扫码登录，按回车继续")
            input()
            page.reload()
            page.wait_for_timeout(3000)
        
        print("✅ 已登录")
        
        # 进入作品分析
        print("\n进入作品分析...")
        safe_goto(page, "https://creator.douyin.com/creator-micro/data-center/content")
        
        # 点击投稿列表
        print("\n点击投稿列表...")
        for _ in range(10):
            try:
                page.click('text=投稿列表', timeout=2000)
                print("  ✓ 已点击")
                break
            except:
                time.sleep(1)
        
        page.wait_for_timeout(4000)
        
        # 导出数据
        print("\n导出数据...")
        export_success = click_export(page)
        
        # 从 Excel 提取 item_id
        print("\n从 Excel 提取视频列表...")
        export_data = parse_export_excel()
        videos = export_data.get("videos", [])
        
        print(f"  总播放: {export_data.get('total_play', 0):,}")
        print(f"  总点赞: {export_data.get('total_like', 0):,}")
        print(f"  视频数: {len(videos)}")
        
        # 抓取观众画像
        print("\n抓取观众画像...")
        results = []
        for i, v in enumerate(videos[:20]):
            print(f"\n[{i+1}/{min(len(videos),20)}] {v['title'][:50]}")
            
            if not v.get('item_id'):
                print("   无 item_id，跳过")
                continue
            
            profile = fetch_audience_for_video(page, v['item_id'], v['title'])
            if profile:
                results.append({
                    "item_id": v['item_id'],
                    "title": v['title'],
                    "audience": profile
                })
                print(f"   性别: {profile['gender']}")
                print(f"   年龄: {profile['age']}")
                print(f"   地域: {profile['region']}")
            else:
                print("   跳过")
        
        browser.close()
    
    # 保存
    output = {
        "fetch_time": time.strftime("%Y-%m-%d %H:%M:%S"),
        "export_data": {
            "success": export_success,
            "total_play": export_data.get('total_play', 0),
            "total_like": export_data.get('total_like', 0),
            "items_count": export_data.get('items_count', 0)
        },
        "total_videos": len(videos),
        "audience_count": len(results),
        "videos": results
    }
    
    with open("douyin_audience_data.json", "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f"\n{'='*50}")
    print(f"✅ 完成！")
    print(f"   导出: {'✓' if export_success else '✗'}")
    print(f"   画像: {len(results)} 条")
    print(f"   文件: douyin_audience_data.json")


if __name__ == '__main__':
    main()
